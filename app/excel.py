import json
import os.path
import shutil

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

shutil.copyfile(f'C:/Users/stepa/PycharmProjects/HM/app/templates/price_template.xlsx', "test.xlsx")
path = "test.xlsx"
wb = openpyxl.load_workbook(path)
sheet_name = wb.sheetnames
sheet = wb[sheet_name[0]]
with open("price.json", "r", encoding='utf-8') as content:
    price = json.load(content)
    'afffre'


def fill_file(data):
    i = 1
    sheet['B9'].value = data["Услуга"]
    sheet['F4'].value = data["ФИО"]
    sheet['F5'].value = data["ФИО"]
    # should be in try-catch block
    sheet['V9'].value = price[f'{data["Услуга"]}']
    sheet['AD9'].value = price[f'{data["Услуга"]}']
    sheet['AD11'].value = "=AD9*0.13/1.13"
    sheet['AD12'].value = price[f'{data["Услуга"]}']
    sheet['F6'].value = data["Исполнитель"]


def form_file(data):
    fill_file(data)
    wb.save(path)
